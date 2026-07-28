"""
add_aba_custos.py — Adiciona aba "Custos Abertos" ao relatório Portoex.

Uso:
    python add_aba_custos.py \
        --dados_s1   /tmp/dados_s1.json \
        --dados_s2   /tmp/dados_s2.json \
        --custos_csv /path/to/custos_adicionais_brudam.csv \
        --transf_csv /path/to/transferencias_brudam.csv \
        --excel      /path/to/Resultado_Operacional.xlsx

Lógica:
- Seleciona minutas com lucro líquido < -100 (prejuízo > R$ 100)
- Para cada minuta: cabeçalho, resumo financeiro e custo detalhado por linha
- TRANSF. separado em AWB (aéreo) e MANIFESTO (rodoviário) quando disponível
- OUTROS expandido com fornecedor + descrição por lançamento (vindos do CSV)
- Sub-linhas ↳ para múltiplas entradas
"""

import json, csv, argparse
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Cores e estilos ───────────────────────────────────────────────────────────
AZ_ESC  = '1F3864'; AZ_MED  = '2E75B6'; VERM    = 'C00000'
VERM_CL = 'FCE4D6'; VERDE   = '1E7145'; CINZA   = 'EBF0FA'
BRANCO  = 'FFFFFF'; AM_CL   = 'FFF2CC'; AM_SUB  = 'FFF9E6'
AZ_CL   = 'DEEAF1'  # azul claro para sub-linha de TRANSF.

thin = Side(style='thin', color='CCCCCC')
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

def sc(ws, r, c, v, bg, bold=False, color='000000', size=9,
       fmt=None, align='left', italic=False, wrap=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill      = PatternFill('solid', fgColor=bg)
    cell.border    = bdr
    cell.font      = Font(name='Arial', bold=bold, color=color, size=size, italic=italic)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if fmt:
        cell.number_format = fmt
    return cell

# ── Parse de valores BR vindos de CSV ────────────────────────────────────────
def _parse_br_csv(parts, start_idx):
    """
    Converte valor BR (ex: '1.674,02') que pode ter sido dividido em duas
    colunas CSV quando a vírgula decimal foi interpretada como separador.
    Ex: ['1.674','02'] → 1674.02   |   ['353','84'] → 353.84
        ['600,00'] → 600.00  (campo não dividiu, contém vírgula intacta)
    """
    raw = parts[start_idx] if start_idx < len(parts) else ''
    if ',' in raw:
        return float(raw.replace('.', '').replace(',', '.'))
    elif start_idx + 1 < len(parts):
        inteiros = parts[start_idx].replace('.', '')
        decimais = parts[start_idx + 1]
        try:
            return float(f'{inteiros}.{decimais}')
        except Exception:
            return 0.0
    else:
        try:
            return float(raw.replace('.', ''))
        except Exception:
            return 0.0

def parse_br(v):
    if not v or str(v).strip() in ('N/D', 'INDEFINIDO', '', '-', '0,00', '0'):
        return 0.0
    v = str(v).replace('%', '').strip().replace('.', '').replace(',', '.')
    try:
        return float(v)
    except Exception:
        return 0.0

# ── Loaders de CSV ────────────────────────────────────────────────────────────
def load_outros_map(csv_path):
    """
    Lê CSV de custos adicionais (OUTROS) exportado do Brudam.
    Formato: minuta,cod,fornecedor,descricao,valor
    """
    outros_map = {}
    try:
        with open(csv_path, newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader)
            for row in reader:
                if len(row) < 5:
                    continue
                minuta     = row[0]
                fornecedor = row[2]
                descricao  = row[3]
                valor      = _parse_br_csv(row, 4)
                outros_map.setdefault(minuta, []).append({
                    'fornecedor': fornecedor,
                    'descricao':  descricao,
                    'valor':      valor
                })
    except FileNotFoundError:
        print(f'⚠️  {csv_path} não encontrado — OUTROS sem detalhes')
    return outros_map

def load_transf_map(csv_path):
    """
    Lê CSV de transferências exportado do Brudam.
    Formato: minuta,tipo(AWB|MANIFESTO),referencia,empresa,valor
    Retorna: {minuta: {'awb': [...], 'manifesto': [...]}}
    """
    transf_map = {}
    try:
        with open(csv_path, newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader)
            for row in reader:
                if len(row) < 5:
                    continue
                minuta     = row[0]
                tipo       = row[1].upper()
                referencia = row[2]
                empresa    = row[3]
                valor      = _parse_br_csv(row, 4)
                entry = {'referencia': referencia, 'empresa': empresa, 'valor': valor}
                d = transf_map.setdefault(minuta, {'awb': [], 'manifesto': []})
                if tipo == 'AWB':
                    d['awb'].append(entry)
                else:
                    d['manifesto'].append(entry)
    except FileNotFoundError:
        print(f'⚠️  {csv_path} não encontrado — TRANSF. exibida como valor único')
    return transf_map

# ── Geração da aba ────────────────────────────────────────────────────────────
def adicionar_aba_custos(excel_path, dados_s1, dados_s2, outros_map, transf_map):
    wb = load_workbook(excel_path)

    if 'Custos Abertos' in wb.sheetnames:
        del wb['Custos Abertos']

    ws = wb.create_sheet('Custos Abertos')
    ws.sheet_view.showGridLines = False

    col_widths = [16, 34, 38, 16, 16, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Coletar minutas negativas (prejuízo > R$ 100)
    all_rows = []
    for sistema, raw in [('PORTOEX', dados_s1), ('PEX', dados_s2)]:
        for row in raw['rows']:
            if parse_br(row[32]) < -100:
                all_rows.append((sistema, row))
    all_rows.sort(key=lambda x: parse_br(x[1][31]))  # pior primeiro

    # Títulos
    ws.merge_cells('A1:G1')
    sc(ws, 1, 1,
       'CUSTOS ABERTOS — MINUTAS COM PREJUÍZO > R$ 100,00',
       AZ_ESC, bold=True, color=BRANCO, size=12, align='center')
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:G2')
    sc(ws, 2, 1,
       f'{len(all_rows)} minutas  |  Dados extraídos do Brudam em {date.today().strftime("%d/%m/%Y")}',
       '2E4057', italic=True, color='BDC3C7', size=9, align='center')
    ws.row_dimensions[2].height = 14

    col_lbl = ['TIPO DE CUSTO', 'RESPONSÁVEL / EMPRESA', 'REFERÊNCIA / DESCRIÇÃO',
               'VALOR (R$)', 'ACUMULADO (R$)', '% DO CUSTO', '% DO FRETE']
    r = 3

    for sistema, row in all_rows:
        minuta  = row[0];  cliente = row[7];  destino = row[12]; servico = row[10]
        frete   = parse_br(row[13])
        c_col   = parse_br(row[17]); resp_col = (row[16] or 'N/D').strip()
        c_ent   = parse_br(row[19]); resp_ent = (row[18] or 'N/D').strip()
        c_tra   = parse_br(row[21]); resp_tra = (row[20] or 'N/D').strip()
        icms    = parse_br(row[22]); difal    = parse_br(row[23])
        desp    = parse_br(row[24]); seg      = parse_br(row[25])
        outros  = parse_br(row[26]); c_total  = parse_br(row[27])
        imposto = parse_br(row[29]); resultado= parse_br(row[30])
        comis   = parse_br(row[31]); lucro    = parse_br(row[32])
        margem  = parse_br(row[33])

        # Dados de transferência detalhada
        td = transf_map.get(minuta, {'awb': [], 'manifesto': []})
        awb_entries   = td['awb']
        manif_entries = td['manifesto']
        tem_detalhe_transf = awb_entries or manif_entries

        # Cabeçalho da minuta
        ws.merge_cells(f'A{r}:G{r}')
        prefixo = f'[{sistema}]  ' if sistema == 'PEX' else ''
        sc(ws, r, 1,
           f'{prefixo}MINUTA {minuta}   |   {cliente}   |   {servico}   |   {destino}',
           AZ_ESC, bold=True, color=BRANCO, size=10)
        ws.row_dimensions[r].height = 20; r += 1

        # Resumo financeiro
        res = (f'Frete: R$ {frete:,.2f}   Custo Total: R$ {c_total:,.2f}   '
               f'Imposto: R$ {imposto:,.2f}   Resultado: R$ {resultado:,.2f}   '
               f'Comissão: R$ {comis:,.2f}   Lucro Líq.: R$ {lucro:,.2f}   '
               f'Margem: {margem:.2f}%')
        cor_bg = VERM_CL if lucro < 0 else 'E2EFDA'
        cor_t  = VERM   if lucro < 0 else VERDE
        ws.merge_cells(f'A{r}:G{r}')
        sc(ws, r, 1, res, cor_bg, bold=True, color=cor_t, size=9, italic=True)
        ws.row_dimensions[r].height = 14; r += 1

        # Cabeçalho de colunas
        for ci, lbl in enumerate(col_lbl, 1):
            sc(ws, r, ci, lbl, AZ_MED, bold=True, color=BRANCO, size=9, align='center')
        ws.row_dimensions[r].height = 16; r += 1

        acum = 0.0

        # ── COLETA ──
        acum += c_col
        bg = VERM_CL if (c_col > frete > 0) else BRANCO
        pct_c = (c_col / c_total * 100) if c_total else 0
        pct_f = (c_col / frete  * 100) if frete  else 0
        sc(ws, r, 1, 'COLETA',  bg, bold=True, size=9, align='center')
        sc(ws, r, 2, resp_col,  bg, size=9)
        sc(ws, r, 3, '',        bg, size=9)
        sc(ws, r, 4, c_col, bg, bold=True, color=VERM if c_col>frete else '000000',
           size=9, fmt='#,##0.00', align='right')
        sc(ws, r, 5, acum, bg, size=9, fmt='#,##0.00', align='right')
        sc(ws, r, 6, pct_c, bg, size=9, fmt='0.0"%"', align='right')
        sc(ws, r, 7, pct_f, bg, size=9, fmt='0.0"%"', align='right')
        ws.row_dimensions[r].height = 14; r += 1

        # ── ENTREGA ──
        acum += c_ent
        bg = VERM_CL if (c_ent > frete > 0) else CINZA
        pct_c = (c_ent / c_total * 100) if c_total else 0
        pct_f = (c_ent / frete  * 100) if frete  else 0
        sc(ws, r, 1, 'ENTREGA', bg, bold=True, size=9, align='center')
        sc(ws, r, 2, resp_ent,  bg, size=9)
        sc(ws, r, 3, '',        bg, size=9)
        sc(ws, r, 4, c_ent, bg, bold=True, color=VERM if c_ent>frete else '000000',
           size=9, fmt='#,##0.00', align='right')
        sc(ws, r, 5, acum, bg, size=9, fmt='#,##0.00', align='right')
        sc(ws, r, 6, pct_c, bg, size=9, fmt='0.0"%"', align='right')
        sc(ws, r, 7, pct_f, bg, size=9, fmt='0.0"%"', align='right')
        ws.row_dimensions[r].height = 14; r += 1

        # ── TRANSFERÊNCIA — AWB e MANIFESTO separados ──
        if c_tra > 0:
            acum += c_tra
            pct_c_tra = (c_tra / c_total * 100) if c_total else 0
            pct_f_tra = (c_tra / frete  * 100) if frete  else 0

            if tem_detalhe_transf:
                # Primeira linha: cabeçalho TRANSF. com total acumulado
                bg_th = BRANCO
                sc(ws, r, 1, 'TRANSF.',      bg_th, bold=True, size=9, align='center')
                sc(ws, r, 2, resp_tra or '—',bg_th, size=9, color='555555', italic=True)
                sc(ws, r, 3, f'Total: {len(awb_entries)} AWB + {len(manif_entries)} Manifesto',
                   bg_th, size=9, italic=True, color='555555')
                sc(ws, r, 4, c_tra, bg_th, bold=True,
                   color=VERM if c_tra>frete else '000000', size=9, fmt='#,##0.00', align='right')
                sc(ws, r, 5, acum,    bg_th, size=9, fmt='#,##0.00', align='right')
                sc(ws, r, 6, pct_c_tra, bg_th, size=9, fmt='0.0"%"', align='right')
                sc(ws, r, 7, pct_f_tra, bg_th, size=9, fmt='0.0"%"', align='right')
                ws.row_dimensions[r].height = 14; r += 1

                # Sub-linhas AWB
                for entry in awb_entries:
                    sc(ws, r, 1, '↳ AWB',           AZ_CL, size=9, align='center', color='1F3864', bold=True)
                    sc(ws, r, 2, entry['empresa'],   AZ_CL, size=9, color='1F3864')
                    sc(ws, r, 3, f'AWB {entry["referencia"]}', AZ_CL, size=9, italic=True, color='1F3864')
                    sc(ws, r, 4, entry['valor'],     AZ_CL, size=9, color='1F3864', fmt='#,##0.00', align='right')
                    sc(ws, r, 5, '', AZ_CL, size=9)
                    sc(ws, r, 6, '', AZ_CL, size=9)
                    sc(ws, r, 7, '', AZ_CL, size=9)
                    ws.row_dimensions[r].height = 14; r += 1

                # Sub-linhas Manifesto
                for entry in manif_entries:
                    sc(ws, r, 1, '↳ MANIF.',        AZ_CL, size=9, align='center', color='1F3864', bold=True)
                    sc(ws, r, 2, entry['empresa'],   AZ_CL, size=9, color='1F3864')
                    sc(ws, r, 3, f'Manifesto {entry["referencia"]}', AZ_CL, size=9, italic=True, color='1F3864')
                    sc(ws, r, 4, entry['valor'],     AZ_CL, size=9, color='1F3864', fmt='#,##0.00', align='right')
                    sc(ws, r, 5, '', AZ_CL, size=9)
                    sc(ws, r, 6, '', AZ_CL, size=9)
                    sc(ws, r, 7, '', AZ_CL, size=9)
                    ws.row_dimensions[r].height = 14; r += 1
            else:
                # Sem detalhe — linha única
                bg = BRANCO
                sc(ws, r, 1, 'TRANSF.',  bg, bold=True, size=9, align='center')
                sc(ws, r, 2, resp_tra or '—', bg, size=9)
                sc(ws, r, 3, '',         bg, size=9)
                sc(ws, r, 4, c_tra, bg, bold=True, color=VERM if c_tra>frete else '000000',
                   size=9, fmt='#,##0.00', align='right')
                sc(ws, r, 5, acum,       bg, size=9, fmt='#,##0.00', align='right')
                sc(ws, r, 6, pct_c_tra,  bg, size=9, fmt='0.0"%"', align='right')
                sc(ws, r, 7, pct_f_tra,  bg, size=9, fmt='0.0"%"', align='right')
                ws.row_dimensions[r].height = 14; r += 1

        # ── Demais custos ──
        outros_custos = [
            ('ICMS/ISS', '—', icms,  AM_CL),
            ('DIFAL',    '—', difal, AM_CL),
            ('DESPACHO', '—', desp,  AM_CL),
            ('SEGURO',   '—', seg,   AM_CL),
        ]
        for li, (tipo, resp, valor, bg_def) in enumerate(outros_custos):
            if valor == 0:
                continue
            acum += valor
            bg = VERM_CL if (valor > frete > 0) else (CINZA if li % 2 == 0 else BRANCO)
            pct_c = (valor / c_total * 100) if c_total else 0
            pct_f = (valor / frete  * 100) if frete  else 0
            sc(ws, r, 1, tipo,  bg, bold=True, size=9, align='center')
            sc(ws, r, 2, resp,  bg, size=9)
            sc(ws, r, 3, '',    bg, size=9)
            sc(ws, r, 4, valor, bg, bold=True,
               color=VERM if valor>frete else '000000', size=9, fmt='#,##0.00', align='right')
            sc(ws, r, 5, acum,  bg, size=9, fmt='#,##0.00', align='right')
            sc(ws, r, 6, pct_c, bg, size=9, fmt='0.0"%"', align='right')
            sc(ws, r, 7, pct_f, bg, size=9, fmt='0.0"%"', align='right')
            ws.row_dimensions[r].height = 14; r += 1

        # ── OUTROS expandido ──
        if outros > 0:
            entries = outros_map.get(minuta, [])
            acum += outros
            pct_c_out = (outros / c_total * 100) if c_total else 0
            pct_f_out = (outros / frete  * 100) if frete  else 0
            bg_out = VERM_CL if outros > frete else AM_CL

            if entries:
                first = entries[0]
                sc(ws, r, 1, 'OUTROS',           bg_out, bold=True, size=9, align='center')
                sc(ws, r, 2, first['fornecedor'], bg_out, size=9, color='7F4000')
                sc(ws, r, 3, first['descricao'],  bg_out, size=9, italic=True, color='7F4000')
                sc(ws, r, 4, first['valor'],      bg_out, bold=True,
                   color=VERM if outros>frete else '7F4000', size=9, fmt='#,##0.00', align='right')
                sc(ws, r, 5, acum,      bg_out, size=9, fmt='#,##0.00', align='right')
                sc(ws, r, 6, pct_c_out, bg_out, size=9, fmt='0.0"%"', align='right')
                sc(ws, r, 7, pct_f_out, bg_out, size=9, fmt='0.0"%"', align='right')
                ws.row_dimensions[r].height = 14; r += 1

                for entry in entries[1:]:
                    sc(ws, r, 1, '↳',               AM_SUB, size=9, align='center', color='7F4000')
                    sc(ws, r, 2, entry['fornecedor'],AM_SUB, size=9, color='7F4000')
                    sc(ws, r, 3, entry['descricao'], AM_SUB, size=9, italic=True, color='7F4000')
                    sc(ws, r, 4, entry['valor'],     AM_SUB, size=9, color='7F4000',
                       fmt='#,##0.00', align='right')
                    sc(ws, r, 5, '', AM_SUB, size=9)
                    sc(ws, r, 6, '', AM_SUB, size=9)
                    sc(ws, r, 7, '', AM_SUB, size=9)
                    ws.row_dimensions[r].height = 14; r += 1
            else:
                sc(ws, r, 1, 'OUTROS',    bg_out, bold=True, size=9, align='center')
                sc(ws, r, 2, '—',         bg_out, size=9)
                sc(ws, r, 3, '(sem detalhes no período)',
                   bg_out, size=9, italic=True, color='999999')
                sc(ws, r, 4, outros,      bg_out, bold=True, size=9, fmt='#,##0.00', align='right')
                sc(ws, r, 5, acum,        bg_out, size=9, fmt='#,##0.00', align='right')
                sc(ws, r, 6, pct_c_out,   bg_out, size=9, fmt='0.0"%"', align='right')
                sc(ws, r, 7, pct_f_out,   bg_out, size=9, fmt='0.0"%"', align='right')
                ws.row_dimensions[r].height = 14; r += 1

        # ── Total custo ──
        for ci in range(1, 8): sc(ws, r, ci, None, AZ_ESC)
        sc(ws, r, 1, 'TOTAL CUSTO', AZ_ESC, bold=True, color=BRANCO, size=9)
        sc(ws, r, 4, c_total, AZ_ESC, bold=True, color=BRANCO, size=9, fmt='#,##0.00', align='right')
        pct_tf = (c_total / frete * 100) if frete else 0
        sc(ws, r, 7, pct_tf, AZ_ESC, bold=True, color=BRANCO, size=9, fmt='0.0"%"', align='right')
        ws.row_dimensions[r].height = 15; r += 1

        # ── Lucro líquido ──
        lbg = VERM_CL if lucro < 0 else 'E2EFDA'
        lt  = VERM   if lucro < 0 else VERDE
        for ci in range(1, 8): sc(ws, r, ci, None, lbg)
        sc(ws, r, 1, 'LUCRO LÍQ.', lbg, bold=True, color=lt, size=9)
        sc(ws, r, 4, lucro, lbg, bold=True, color=lt, size=9, fmt='#,##0.00', align='right')
        pct_lf = (lucro / frete * 100) if frete else 0
        sc(ws, r, 7, pct_lf, lbg, bold=True, color=lt, size=9, fmt='0.0"%"', align='right')
        ws.row_dimensions[r].height = 15; r += 2  # linha em branco entre minutas

    wb.save(excel_path)
    print(f'✅  Aba "Custos Abertos" adicionada → {excel_path}')
    print(f'    Minutas: {len(all_rows)} | Linhas Excel: {r}')


# ── CLI ───────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    ap = argparse.ArgumentParser(description='Adiciona aba Custos Abertos ao relatório Portoex')
    ap.add_argument('--dados_s1',   required=True, help='JSON do PORTOEX (dados_s1.json)')
    ap.add_argument('--dados_s2',   required=True, help='JSON da PEX (dados_s2.json)')
    ap.add_argument('--custos_csv', required=True, help='CSV de custos adicionais (OUTROS) do Brudam')
    ap.add_argument('--transf_csv', required=True, help='CSV de transferências (AWB/Manifesto) do Brudam')
    ap.add_argument('--excel',      required=True, help='Caminho do Excel a ser atualizado')
    args = ap.parse_args()

    with open(args.dados_s1) as f: s1 = json.load(f)
    with open(args.dados_s2) as f: s2 = json.load(f)

    outros_map = load_outros_map(args.custos_csv)
    transf_map = load_transf_map(args.transf_csv)
    adicionar_aba_custos(args.excel, s1, s2, outros_map, transf_map)
