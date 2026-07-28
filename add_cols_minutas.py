#!/usr/bin/env python3
"""
add_cols_minutas.py — Adiciona Valor NF, Vol e Peso (kg) à aba Minutas do relatório.

Uso:
    python3 add_cols_minutas.py \
        --dados_s1  /tmp/merged_s1.json \
        --dados_s2  /tmp/merged_s2.json \
        --excel     /path/to/Resultado_Operacional.xlsx

Lógica:
- Lê os JSONs merged (merged_s1.json, merged_s2.json) e constrói lookup MINUTA → {vnf, vol, peso}
- Abre o Excel, encontra a aba "Minutas", adiciona 3 colunas ao final
- Preserva o estilo (fonte, borda, preenchimento) das células adjacentes
"""

import json, copy, argparse, os
import openpyxl
from openpyxl.utils import get_column_letter

def parse_br(v):
    if not v or str(v).strip() in ('N/D', 'INDEFINIDO', '', '-', '0,00', '0'): return 0.0
    v = str(v).replace('%', '').strip().replace('.', '').replace(',', '.')
    try: return float(v)
    except: return 0.0

def build_lookup(*json_paths):
    lookup = {}
    for path in json_paths:
        if not os.path.exists(path): continue
        with open(path) as f: raw = json.load(f)
        hdrs = raw['headers']
        try:
            idx_min  = hdrs.index('MINUTA')
            idx_vnf  = hdrs.index('VALOR NF')
            idx_vol  = hdrs.index('VOLUMES')
            idx_peso = hdrs.index('PESO')
        except ValueError as e:
            print(f"  Aviso: coluna não encontrada em {path}: {e}")
            continue
        for row in raw['rows']:
            minuta = row[idx_min].strip()
            if minuta:
                lookup[minuta] = {
                    'vnf':     parse_br(row[idx_vnf]),
                    'volumes': parse_br(row[idx_vol]),
                    'peso':    parse_br(row[idx_peso]),
                }
    return lookup

def clone_style(src, dst):
    if src.font:      dst.font      = copy.copy(src.font)
    if src.fill:      dst.fill      = copy.copy(src.fill)
    if src.border:    dst.border    = copy.copy(src.border)
    if src.alignment: dst.alignment = copy.copy(src.alignment)

def adicionar_colunas_minutas(excel_path, lookup):
    wb = openpyxl.load_workbook(excel_path)
    if 'Minutas' not in wb.sheetnames:
        print("Aba 'Minutas' não encontrada — pulando.")
        return

    ws = wb['Minutas']
    max_col = ws.max_column
    hdr_row = 2

    # Encontrar coluna Minuta
    col_minuta = None
    for col in range(1, max_col + 1):
        v = ws.cell(hdr_row, col).value
        if v and 'Minuta' in str(v):
            col_minuta = col
            break
    if not col_minuta:
        print("Coluna 'Minuta' não encontrada na aba Minutas — pulando.")
        return

    new_cols = [
        (max_col + 1, 'Valor NF (R$)', 'R$ #,##0.00', 14),
        (max_col + 2, 'Vol',           '#,##0',         8),
        (max_col + 3, 'Peso (kg)',     '#,##0.0',       11),
    ]

    # Larguras
    for col_idx, _, _, width in new_cols:
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Cabeçalhos
    ref_hdr = ws.cell(hdr_row, max_col)
    for col_idx, label, _, _ in new_cols:
        c = ws.cell(hdr_row, col_idx, value=label)
        clone_style(ref_hdr, c)

    # Estender merge da linha 1 (título)
    for merge in list(ws.merged_cells.ranges):
        if merge.min_row == 1:
            ws.merged_cells.remove(merge)
            ws.merge_cells(
                start_row=1, start_column=merge.min_col,
                end_row=1, end_column=max_col + len(new_cols)
            )
            break

    # Preencher dados
    hits = 0
    for row in range(hdr_row + 1, ws.max_row + 1):
        minuta_val = str(ws.cell(row, col_minuta).value or '').strip()
        ref_data = ws.cell(row, max_col)
        for col_idx, _, fmt, _ in new_cols:
            dc = ws.cell(row, col_idx)
            clone_style(ref_data, dc)
            dc.number_format = fmt
        if minuta_val and minuta_val.isdigit() and minuta_val in lookup:
            d = lookup[minuta_val]
            ws.cell(row, new_cols[0][0]).value = d['vnf'] or None
            ws.cell(row, new_cols[1][0]).value = int(d['volumes']) if d['volumes'] else None
            ws.cell(row, new_cols[2][0]).value = d['peso'] if d['peso'] else None
            hits += 1

    # Atualizar auto_filter
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(max_col + len(new_cols))}{ws.max_row - 1}"

    wb.save(excel_path)
    print(f"✅ Colunas adicionadas: {hits} minutas preenchidas com Valor NF / Vol / Peso")

if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--dados_s1', default='/tmp/merged_s1.json')
    ap.add_argument('--dados_s2', default='/tmp/merged_s2.json')
    ap.add_argument('--excel', required=True)
    args = ap.parse_args()

    lookup = build_lookup(args.dados_s1, args.dados_s2)
    print(f"Lookup: {len(lookup)} minutas")
    adicionar_colunas_minutas(args.excel, lookup)
